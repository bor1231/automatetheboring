import openpyxl
wb = openpyxl.load_workbook('example3.xlsx')
print(type(wb))
print('the sheet names are: ')
print(wb.sheetnames)
sheet = wb['Sheet3']
print(sheet.title)
another_sheet = wb.active
print('The active sheet is: ' + str(another_sheet))
#getting cell values
print('The value of A1 is : ' + str(another_sheet['A1'].value))
g = another_sheet['B1']
s = another_sheet
print(f'The value of cell with row {g.row} and column {g.column} is : ' + str(another_sheet['B1'].value) + f'. That was for {g.coordinate}')
print(s.cell(row=1, column=2))
for i in range(1,8,2):
    print(i, s.cell(row=i, column=2).value)
print(f' the size of the workbook is {s.max_row} x {s.max_column}')

#slicing the data

slice = s['A1':'C3']
for row in slice:
    for j in row:
        print(j.coordinate, j.value)
    print('---END of ROW---')

#data can be arranged in columns or rows:
some_data  = list (another_sheet.columns)
print('organized by colums:', some_data[1])
for cell in some_data[1]:
    print(cell.value)
