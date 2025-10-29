import openpyxl

#to open and change the sheet name and save the file to keep the changes

wb = openpyxl.load_workbook('sensor_data_output.xlsx')
print(type(wb))
print('the sheet names are: ' + str(wb.sheetnames))
sheet = wb.active
sheet.title = 'Peltier'
print('now the sheet name are: ' + str(wb.sheetnames))
wb.save('sensor_data_output.xlsx')
print('the file has been saved.')